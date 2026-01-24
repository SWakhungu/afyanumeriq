import os
import re
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from api.isms_models import Clause, Control, SoAEntry
from api.models import ComplianceClause

CONTROL_CODE_PATTERN = re.compile(r"^A\.\d+(\.\d+)?$")


class Command(BaseCommand):
    help = "Import ISO 27001 Clauses and Annex A Controls from Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Path to the Excel file containing Clauses and Controls",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = options["file"]

        if not os.path.exists(path):
            self.stdout.write(self.style.ERROR(f"File not found: {path}"))
            return

        wb = load_workbook(path)

        # --------------------------------------------------------
#                IMPORT ISO 27001 CLAUSES
# --------------------------------------------------------
        if "Clauses" in wb.sheetnames:
            sheet = wb["Clauses"]

            ComplianceClause.objects.filter(standard="iso-27001").delete()
            count = 0

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header

                clause_code = str(row[0]).strip() if row[0] else None
                description = str(row[1]).strip() if row[1] else ""

                if not clause_code:
                    continue

                ComplianceClause.objects.update_or_create(
                    standard="iso-27001",
                    clause_number=clause_code,
                    defaults={
                        "short_description": description[:200],
                        "description": description,
                        "status": "NI",
                        "owner": "Unassigned",
                    },
                )
                count += 1

            self.stdout.write(self.style.SUCCESS(f"✔ ISO 27001 Clauses imported: {count}"))

        # ========================================================
        #                   IMPORT CONTROLS
        # ========================================================
        if "Controls" in wb.sheetnames:
            sheet = wb["Controls"]

            Control.objects.filter(standard="iso-27001").delete()
            SoAEntry.objects.filter(standard="iso-27001").delete()

            control_count = 0
            current_group = ""

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header

                raw_code = str(row[0]).strip() if row[0] else ""
                title = str(row[1]).strip() if row[1] else ""

                if not raw_code:
                    continue

                # Detect group headers
                if raw_code.startswith("A.") and "Controls" in raw_code:
                    current_group = raw_code
                    continue

                if not CONTROL_CODE_PATTERN.match(raw_code):
                    continue

                ctrl = Control.objects.create(
                    code=raw_code,
                    title=title,
                    description="",
                    group=current_group,
                    standard="iso-27001",
                )

                SoAEntry.objects.create(
                    control=ctrl,
                    standard="iso-27001",
                    applicable=True,
                    status="Not implemented",
                    justification="",
                    evidence_notes="",
                )

                control_count += 1

            self.stdout.write(
                self.style.SUCCESS(f"✔ Controls imported: {control_count}")
            )
        else:
            self.stdout.write(self.style.WARNING("No 'Controls' sheet found."))

        self.stdout.write(self.style.SUCCESS("✅ ISO 27001 import complete."))
